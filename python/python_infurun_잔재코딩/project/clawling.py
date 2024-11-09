"""

크롤링 프로젝트 수행 

"""

import requests
from bs4 import BeautifulSoup
import openpyxl
import re

## 단계별로 진행

"""
1. 타이틀 가져와서 출력하기 
2. 타이틀 인덱스 출력하기 
3. 댓글 수 출력하기 
4. 링크 출력하기
"""

url = "https://davelee-fun.github.io/trial/board/news.html"

res = requests.get(url)

data = BeautifulSoup(res.content, "html.parser")

# 데이터의 각 행을 의미한다.
title = data.select("div.list_item")


# ## 0-9까지의 숫자가 하나이상 반복되고, 뒤에 공백이 있던지 없던지의 패턴을 찾는다.
# text_pattern = re.compile("[0-9]+\s*\t\f")


"""
excel 파일 조작 
"""
import openpyxl
from openpyxl.styles import Alignment, Font

excel_file = openpyxl.Workbook()

# 활성화된 시트를 가져오기
sheet = excel_file.active

# 데이터 추가하기
sheet.title = "게시글"


wrap_alignment = Alignment(wrap_text=True)
bold_font = Font(bold=True)
header_font = Font(bold=True, size=14)
center_alignment = Alignment(horizontal="center", vertical="center")
index = 0

header = ["순번", "게시글 제목", "댓글"]

sheet.append(header)

## sheet[1]은 헤더행을 의미한다.
for row in sheet.iter_rows():
    for cell in row:
        cell.font = header_font
        cell.alignment = center_alignment
        print(f"셀 위치: {cell.coordinate}, 셀 값: {cell.value}")


sheet.column_dimensions["A"].width = 50
sheet.column_dimensions["B"].width = 20
sheet.column_dimensions["C"].width = 100


"""
게시글의 댓글 내용을 출력하려고 하면, 게시글 행의 주소에서 다시 크롤링을 수행해주어야 한다. 

"""


for items in title[:10]:
    title = items.select_one("span.subject_fixed")
    comment = items.select_one("a.list_reply")
    link = items.select_one("a.list_subject")
    data = {}

    if title != None:
        link_url = "https://davelee-fun.github.io/trial/board/" + link["href"]
        res_title = requests.get(link_url)
        soup_title = BeautifulSoup(res_title.content, "html.parser")
        post_comment = soup_title.select("div.comment_view")
        index += 1  # 인덱스 추가
        data["index"] = index  # 인덱스 추가
        data["title"] = title.get_text().strip()
        # sheet.cell(row=index+1, column=2).hyperlink=link_url

        comments_text = []
        print(index, title.get_text().strip())
        print(f"data 딕셔너리{data}")
        for idx, reply in enumerate(post_comment):
            # comment_text는 문자열로 출려된다.
            comment_text = reply.get_text().strip().replace("\n", "").replace("\t", "")
            print(f"타입 체크 {type(comment_text)}")
            comments_text.append(f"{idx+1}. {comment_text}\n")
            print(f"댓글 리스트 {comments_text}")
            # excel에는 리스트 형태의 데이터를 추가할 수 없으므로, 문자열로 만들어주어야 한다.
            data["comment"] = "".join(comments_text).strip()
            # data["comment"] = comment_text
            print(f"data {data['comment']}")
            print("ㄴ", reply.get_text().strip().replace("\n", "").replace("\t", ""))
        sheet.append(list(data.values()))  # dict.values()를 list로 변환하여 append
        sheet.cell(row=index + 1, column=2).hyperlink = link_url
        sheet.cell(row=index + 1, column=1).alignment = center_alignment
        sheet.cell(row=index + 1, column=1).font = Font(bold=True, size=10)
excel_file.save("./게시글_크롤링_결과.xlsx")
excel_file.close()
